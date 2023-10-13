from pywinauto.application import Application
from pywinauto import keyboard
import pandas as pd
import time
from pywinauto import findwindows
from openpyxl import Workbook
from openpyxl import load_workbook
import subprocess
from PIL import ImageGrab


class mybillbookbase:


    def __init__(self):
        self.app = (Application(backend='uia').start(r'C:\Users\ponraj.kasirajan\AppData\Local\Programs\mybillbook-desktop-app\MyBillbook.exe').connect(title='My BillBook | Simple Billing and Inventory Software', timeout=30))
        pass

    def clickingcreatesalesinvoice(self):
        csibutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Create Sales Invoice",control_type="Text").wrapper_object()
        csibutton.click_input()
        pass
    def clickingdropdown(self):
        dropdownbtn = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="open create dropdown",control_type="Image").wrapper_object()
        dropdownbtn.click_input()
        pass
    def clickingparties(self):
        partybutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Create Party", control_type="Text").wrapper_object()
        partybutton.click_input()
        pass
    def partynameinput(self, name):
        partynameinputbox = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Party Name*",control_type="Edit").wrapper_object()
        partynameinputbox.set_text(name)
        pass
    def mobilenumber(self,mn):
        mobilenumberinputbox = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Mobile Number",control_type="Edit").wrapper_object()
        mobilenumberinputbox.set_text(mn)
        pass
    def emailinputbox(self,email):
        emailinputbox = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Email", control_type="Edit").wrapper_object()
        emailinputbox.set_text(email)
        pass
    def openingbalancefield(self,amount):
        openningbalance = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="0", control_type="Edit").wrapper_object()
        openningbalance.set_text(amount)
        pass

    def partycategory(self):
        partycategory = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Select Category", auto_id="dropdownMenuButton", control_type="Group").wrapper_object()
        partycategory.click_input()
        pass
    def createcategory(self):
        createcategory = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="+ Create Category", control_type="Text", top_level_only=False).wrapper_object()
        createcategory.click_input()
        pass

    def createcategoryname(self,categoryname):
        newcategoryname = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter category name", control_type="Edit").wrapper_object()
        newcategoryname.set_text(categoryname)
        pass

    def savecategoryname(self):
        savenewcategory = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Save", control_type="Button", found_index=1).wrapper_object()
        savenewcategory.click_input()
        pass
    def gstnumberinputbox(self,gstnumber):
        GSTinputbox = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="ex: 29XXXXX9438X1XX", control_type="Edit").wrapper_object()
        GSTinputbox.set_text(gstnumber)
        pass

    def pannumberinputbox(self,pannumber):
        PANnumberinputbox = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter party PAN Number", control_type="Edit").wrapper_object()
        PANnumberinputbox.set_text(pannumber)
        pass

    def billingaddresspopup(self):
        enterbillingaddresspopup = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter billing address", control_type="Edit").wrapper_object()
        enterbillingaddresspopup.click_input()
        pass
    def billingaddress_streetname(self,address):
        try: billingaddress_streetname = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter Street Address", auto_id="street_addr", control_type="Edit").wrapper_object()
        except findwindows.ElementNotFoundError:
            if self.app:
                self.app.kill()
        except Exception as e:
            print(f"An error occurred: {e}")
            if self.app:
                self.app.kill()

        billingaddress_streetname.set_text(address)
        pass

    def billingaddress_state(self,state):
        billingaddress_statesi = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter State", auto_id="searchInput", control_type="Edit").wrapper_object()
        billingaddress_statesi.set_text(state)
        pass

    def billingaddress_pincode(self, pincode):
        billingaddress_pincode = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter pin code", auto_id="pincode", control_type="Edit").wrapper_object()
        billingaddress_pincode.set_text(pincode)
        pass

    def billingaddress_city(self, city):
        billingaddress_city = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter City", auto_id="city", control_type="Edit").wrapper_object()
        billingaddress_city.set_text(city)
        pass
    def billingaddress_savebutton(self):
        billingaddress_savebutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Save", control_type="Button", found_index=1).wrapper_object()
        billingaddress_savebutton.click_input()
        pass

    def categorysuccessmsgcloseicon(self):
        categorypopupcloseicon = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="close", control_type="Text", found_index=0).wrapper_object()
        categorypopupcloseicon.click_input()
        pass
    def partysavebutton(self):
        partysavebutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Save", control_type="Button").wrapper_object()
        partysavebutton.click_input()
        pass

    def menu_parties(self):
        menu_parties = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Parties", control_type="Hyperlink").wrapper_object()
        menu_parties.click_input()
        pass

    def getnameandverifyname(self,name):
        getname = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title=name, control_type="Text").wrapper_object()
        textofname = getname.window_text()
        if textofname==name:
            print("Record saved and its reflected in screen")

        pass

    def getandverifycategory(self, categoryname):
        getcategory = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title=categoryname, control_type="DataItem").wrapper_object()
        textofcategory = getcategory.window_text()
        if textofcategory == categoryname:
            print("Record saved and its reflected in screen")

        pass
    def searchpartbyname(self, name):
        searchpartybyname = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Search party by name or number", control_type="Edit").wrapper_object()
        searchpartybyname.set_text(name)
    pass

    def wrong_element(self):
        menu_parties2 = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Parties",control_type="Hyperlink").wrapper_object()
        menu_parties2.click_input()
        pass
    def closingwindow(self):
        keyboard.send_keys('%{F4}')
        pass

    def reportmenu(self):
        reportmenu = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Reports", control_type="Hyperlink").wrapper_object()
        reportmenu.click_input()
        pass
    def printidentifiers(self):
        self.app.MyBillBookSimpleBillingAndInventorySoftware.print_control_identifiers()
        pass

    def balancesheet(self):
        balancesheet = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Balance Sheet", control_type="Text").wrapper_object()
        balancesheet.click_input()
        pass
    def additemincurrentasset(self):
        additemincurrentasset = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="+ Add New Entry", control_type="Text",found_index=3).wrapper_object()
        additemincurrentasset.click_input()
        pass
    def ledgername(self):
        additemincurrentasset = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Ledger Name", control_type="Text").wrapper_object()
        additemincurrentasset.set_text('ponraj')
        pass

    def downloadexcelsheetfrombalancesheet(self):
        downloadexcelsheetfrombalancesheet = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Excel Download", control_type="Button").wrapper_object()
        downloadexcelsheetfrombalancesheet.click_input()
        pass
    def readdatafromexcelafterdownload(self):
        df = pd.read_excel(r'C:\Users\ponraj.kasirajan\Downloads\jp_manageme_balance_sheet.xlsx')
        #excelsheetopen = Application(backend='uia').start(r'C:\Users\ponraj.kasirajan\Downloads\jp_balance_sheet.xlsx')
        #print(df)
        pass

    def readdatafromexcelafterdownloadwp(self):
        df = pd.read_excel(r'C:\Users\ponraj.kasirajan\jp_balance_sheet.xlsx')
        #excelsheetopen = Application(backend='uia').start(r'C:\Users\ponraj.kasirajan\Downloads\jp_balance_sheet.xlsx')
        #print(df)
        pass
    def waittillload(self):
        time.sleep(5)
        pass
    def selectingpartyfordelete(self):
        username = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Fusion Technology", control_type="Text").wrapper_object()
        username.click_input()
        pass

    def selectinginvoiceuser(self):
        username = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Biotech Industries", control_type="Text").wrapper_object()
        username.click_input()
        pass

    def clickingdeletebutton(self):
        clickingdeletebutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Delete", control_type="Button").wrapper_object()
        clickingdeletebutton.click_input()
        pass

    def clickingyes_deletebutton(self):
        clickingyes_deletebutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Yes, Delete", control_type="Button").wrapper_object()
        clickingyes_deletebutton.click_input()
        pass

    def csi(self):
        csibutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Create Sales Invoice", control_type="Text").wrapper_object()
        csibutton.click_input()
        pass

    def shift_y(self):
        keyboard.send_keys('+{Y}')
        pass
    def addparty(self):
        csi_addparty = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="+ Add Party", control_type="Text").wrapper_object()
        csi_addparty.click_input()
        pass
    def csi_selectingparty(self):
        csi_selectingparty = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Biotech Industries", control_type="Text").wrapper_object()
        csi_selectingparty.click_input()
        pass
    def csi_selectingdate(self):
        csi_selectingdate = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="20 Sep 2023", control_type="Group").wrapper_object()
        csi_selectingdate.click_input()
        pass
    def csi_modifyingdate(self):
        csi_modifyingdate = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="13", control_type="DataItem").wrapper_object()
        csi_modifyingdate.click_input()
        pass
    def csi_createitem(self):
        csi_createitem = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="+ Create Item", control_type="Text").wrapper_object()
        csi_createitem.click_input()
        pass

    def csi_serviceradiobutton(self):
        csi_serviceradiobutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Service", auto_id="item_type2", control_type="RadioButton").wrapper_object()
        csi_serviceradiobutton.click_input()
        pass

    def csi_itemname(self):
        csi_createname = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Enter name", auto_id="firstItemInput", control_type="Edit").wrapper_object()
        csi_createname.set_text('documentation')
        pass



    def csi_salesprice(self):
        csi_salesprice = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="0", control_type="Edit").wrapper_object()
        csi_salesprice.set_text('3000')
        pass

    def csi_savebutton(self):
        csi_savebutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Save", control_type="Button",found_index=1).wrapper_object()
        csi_savebutton.click_input()
        pass
    def csi_addfromsampleitem(self):
        csi_addfromsampleitem = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="+ Add", control_type="DataItem", found_index=1).wrapper_object()
        csi_addfromsampleitem.click_input()
        pass
    def csi_donebutton(self):
        csi_donebutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Done", control_type="Button").wrapper_object()
        csi_donebutton.click_input()
        pass

    def csi_markasfullypaidcheckbox(self):
        csi_fullypaid = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Mark as fully paid", control_type="Text").wrapper_object()
        csi_fullypaid
        pass

    def csi_formsavebutton(self):
        csi_formsavebutton2 = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Save", auto_id="submit-button", control_type="Button").wrapper_object()
        csi_formsavebutton2.click_input()
        pass

    def csi_scrolldownpage(self):
        list_control = self.app.window(title = 'My BillBook | Simple Billing and Inventory Software')
        scroll_amount = 1
        list_control.scroll(direction='down',amount =scroll_amount )
        pass

    def csi_scrolluppage(self):
        list_control = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        scroll_amount = 2
        list_control.scroll(direction='up',amount =scroll_amount )
        pass
    def csi_downkey(self):
        self.app=keyboard.send_keys('{VK_DOWN}')
        pass

    def closingthe_errormsg(self):
        closingtheerrormsg2 = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="close", control_type="Text", found_index=1).wrapper_object()
        closingtheerrormsg2.click_input()
        pass

    def bulkimportbutton(self):
        bulkimportbutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Bulk Import Parties", control_type="Button").wrapper_object()
        bulkimportbutton.click_input()
        pass
    def uploadpartybutton(self):
                uploadpartybutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Upload Parties", control_type="Text").wrapper_object()
                uploadpartybutton.click_input()
    pass

    def biotechindustries(self):
        bio= self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Biotech Industries", control_type="Text").wrapper_object()
        bio.click_input()
        pass
    def transactionbutton(self):
     trasactionbutton = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window(title="Transactions", control_type="Text").wrapper_object()
     trasactionbutton.click_input()
     pass
    def listofdata(self):
        parent_element = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        data_items = parent_element.descendants(control_type="DataItem")
        workbook = load_workbook("transactions_records.xlsx")
        #workbook = Workbook()
        sheet = workbook["Sheet"]

        sheet = workbook.active

        row_index = 2
        column_index = 1

        for data_item in data_items:
            data = data_item.texts()
            for value in data:
                sheet.cell(row=row_index, column=column_index, value=value)
                column_index += 1
                if column_index % 5 == 1:  # Start a new column after every 4 data items
                    row_index += 1
                    column_index = 1

       # for row_index, data_item in enumerate(data_items, start=1):
        #    data = data_item.texts()
         #   for column_index, value in enumerate(data, start=1):
          #      sheet.cell(row=row_index, column=column_index, value=value)

        filepath = "transactions_records.xlsx"
        workbook.save(filepath )
        time.sleep(4)
        subprocess.Popen([filepath], shell=True)
        time.sleep(15)
        windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        time.sleep(1)
        screenshot = windowforscreenshot.capture_as_image()
        screenshot.save("screenshot\gettingdatafromexcelsheet.png")

        self.app.kill()


        #subprocess.Popen([filepath ], shell=True)

        #list_view_element = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window()

        #items = list_view_element.get_items()
        #items_list = items.split("\n")
       # print(items)

       # for data_item in data_items:
            #print(data_item.texts())


        pass

    def listofheader(self):
        parent_element = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        data_items = parent_element.descendants(control_type="Header")
        workbook = Workbook()
        sheet = workbook.active

        row_index = 1
        column_index = 1

        for data_item in data_items:
            data = data_item.texts()
            for value in data:
                sheet.cell(row=row_index, column=column_index, value=value)
                column_index += 1
                if column_index % 5 == 1:  # Start a new column after every 4 data items
                    row_index += 1
                    column_index = 1

       # for row_index, data_item in enumerate(data_items, start=1):
        #    data = data_item.texts()
         #   for column_index, value in enumerate(data, start=1):
          #      sheet.cell(row=row_index, column=column_index, value=value)

        workbook.save("transactions_records.xlsx")

        #list_view_element = self.app.MyBillBookSimpleBillingAndInventorySoftware.child_window()

        #items = list_view_element.get_items()
        #items_list = items.split("\n")
       # print(items)

       # for data_item in data_items:
            #print(data_item.texts())


        pass


    def screenshot1(self):

      windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
      time.sleep(1)
      screenshot = windowforscreenshot.capture_as_image()
      screenshot.save("screenshot\creataparty.png")
    pass

    def screenshot2(self):

      windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
      time.sleep(1)
      screenshot = windowforscreenshot.capture_as_image()
      screenshot.save("screenshot\duplicateparty.png")
    pass

    def screenshot3(self):

        windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        time.sleep(1)
        screenshot = windowforscreenshot.capture_as_image()
        screenshot.save("screenshot\deleteparty.png")

    pass

    def screenshot4(self):

        windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        time.sleep(1)
        screenshot = windowforscreenshot.capture_as_image()
        screenshot.save("screenshot\creatinginvoice.png")

    pass

    def screenshot5(self):

        windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        time.sleep(1)
        screenshot = windowforscreenshot.capture_as_image()
        screenshot.save("screenshot\gettingdatafromexcelsheet.png")

    pass

    def screenshot6(self):

        windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        time.sleep(1)
        screenshot = windowforscreenshot.capture_as_image()
        screenshot.save("screenshot\deleteapartywithsalesinvoice.png")

    pass

    def screenshot6(self):

        windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        time.sleep(1)
        screenshot = windowforscreenshot.capture_as_image()
        screenshot.save("screenshot\deleteapartywithsalesinvoice.png")

    pass

    def screenshot7(self):

        windowforscreenshot = self.app.window(title='My BillBook | Simple Billing and Inventory Software')
        time.sleep(1)
        screenshot = windowforscreenshot.capture_as_image()
        screenshot.save("screenshot\dowloadingexcelsheet.png")

    pass

